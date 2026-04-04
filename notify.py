import os
from urllib.parse import urlparse
import requests
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Reads Telegram bot token and chat ID from a file
def read_telegram_info(tginfo_path='tginfo.txt'):
    """Read Telegram bot token and chat ID from a text file."""
    with open(tginfo_path, 'r') as f:
        lines = f.read().splitlines()
        token = lines[0].strip()
        chat_id = lines[1].strip()
    return token, chat_id

# Extracts the site name from a given URL for labeling messages
def extract_site_name(url):
    """Extract the site name from a URL for labeling."""
    # Custom handling for specific foxbet URLs
    if 'foxbet.gr/316026/to-dunato-simeio' in url:
        return 'foxbet-to_dynato'
    elif 'foxbet.gr/316014/to-stantar' in url:
        return 'foxbet-to_stantar'
    # Custom handling for nostrabet URLs
    elif 'nostrabet.com/en/bet-of-the-day' in url:
        return 'nostra_bet_of_the_day'
    elif 'nostrabet.com/en/banker-of-the-day' in url:
        return 'notra_banker_of_the_day'
    # Custom handling for kingbet URLs
    elif 'kingbet.com.cy/to-dynato-simeio-imeras' in url:
        return 'kingbet-to_dynato'
    elif 'kingbet.com.cy/favori-imeras' in url:
        return 'kingbet-to_favori'
    elif 'kingbet.com.cy/to-stantar-tis-imeras' in url:
        return 'kingbet-to_stantar'

    # Default behavior for other URLs
    netloc = urlparse(url).netloc
    if netloc.startswith('www.'):
        netloc = netloc[4:]
    site_name = netloc.split('.')[0]
    return site_name if site_name else 'unknown'

# Returns the monthly log file name for tips
def get_monthly_log_file(base_name="tips_log"):
    """Generate the monthly log file name for tips."""
    month_str = datetime.today().strftime("%Y_%m")
    return f"{base_name}_{month_str}.xlsx"

# Logs a tip to the monthly Excel file, adding headers if new
def log_tip_to_xlsx(xlsx_file, url, tip):
    """Log a tip to the monthly Excel file, adding headers if the file is new."""
    now = datetime.now()
    date_str = now.strftime("%d/%m/%Y")
    time_str = now.strftime("%H:%M")
    tipster = extract_site_name(url)
    if not os.path.exists(xlsx_file):
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Time", "Tipster", "Tip"])
        wb.save(xlsx_file)
    wb = load_workbook(xlsx_file)
    ws = wb.active
    ws.append([date_str, time_str, tipster, tip])
    wb.save(xlsx_file)

# Normalize message text for Telegram: trim each line and drop empty lines.
def normalize_message_text(message, lines_to_trim):
    cleaned_lines = [line.strip() for line in message.splitlines() if line.strip()]
    return "\n".join(cleaned_lines[:lines_to_trim])

# Sends a message to a Telegram chat using the bot API and logs the tip
def send_message(message, url=None, tginfo_path='tginfo.txt', lines_to_trim=10):
    """Send a message to Telegram and log the tip to the monthly Excel file."""
    trimmed_message = normalize_message_text(message, lines_to_trim)
    # Only send if the tip content is not empty (excluding site name)
    if not trimmed_message.strip():
        print("⚠️ Empty message, not sending to Telegram.")
        return False
    token, chat_id = read_telegram_info(tginfo_path)
    if url:
        site_name = extract_site_name(url)
        # If the trimmed_message is empty after stripping, do not send even with site name
        if not trimmed_message.strip():
            print(f"⚠️ Empty tip for [{site_name}], not sending to Telegram.")
            return False
        trimmed_message = f"[{site_name}] {trimmed_message}"
    url_api = f'https://api.telegram.org/bot{token}/sendMessage'
    payload = {'chat_id': chat_id, 'text': trimmed_message}
    response = requests.post(url_api, data=payload)
    print("Telegram response:", response.text)
    if url:
        log_tip_to_xlsx(get_monthly_log_file(), url, trimmed_message)
    return response.ok

def main():
    # Entry point for notification handling
    pass

if __name__ == "__main__":
    main()